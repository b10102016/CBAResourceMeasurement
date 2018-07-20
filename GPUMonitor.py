# -*- coding: utf-8 -*-

import time

import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.chart import LineChart, Reference, Series
import chromeDino
import threading
def dumpsys_gfxinfo_framestats(ip,packageName,startIntent):

    # This is for Android 6.0 above
    # Author：John Hao
    # 測試流暢度平均繪製時長的小腳本
    # 功能：滑動當前頁面兩次，間隔1秒，收集gfxinfo並製表
    # titlename：表示要測試的模塊和最後文件保存的名字
    # |-- 收集gfxinfo framestats詳細幀數據信息
    # |-- 用第三方模塊openpyxl處理收集到的excel數據
    # |-- 輸出2秒內的120幀每一幀的繪製時間
    # |-- 輸出平均繪製時間

    


    titlelist = [' UI 線程中發生的工作使其無法及時響應垂直同步信號','應用處理輸入事件所花的時間（ >2 毫秒錶示處理輸入事件時間長）',
                    '正在運行的所有動畫（ObjectAnimator、ViewPropertyAnimator 和通用轉換）所需的時間',
                    '完成佈局和測量階段所需的時間','對樹中的所有視圖調用 View.draw() 所需的時間',
                    '約 > 0.4 毫秒錶示繪製了大量必須上傳到 GPU 的新位圖','GPU 工作量','處理此幀所花的總時間','達標線']
    subtitlelist = ['IntendedVsync','HandleInputStart','AnimationStart','PerformTraversalsStart',
                    'DrawStart','SyncStart','IssueDrawCommandsStart','FrameCompleted','Avg']

    # 要測試測模塊名，最後文件會以該名稱命名
    titlename = "Feed"
    print "Starting"
    os.system("adb -s "+ip+" shell am start "+packageName+"/"+startIntent) 
    for j in range(1,6):
        time.sleep(1)
        print "開始執行第" + str(j) + "遍"

        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        valueofwidth = 16
        ws.column_dimensions["A"].width = valueofwidth
        ws.column_dimensions["B"].width = valueofwidth
        ws.column_dimensions["C"].width = valueofwidth
        ws.column_dimensions["D"].width = valueofwidth
        ws.column_dimensions["E"].width = valueofwidth
        ws.column_dimensions["F"].width = valueofwidth
        ws.column_dimensions["G"].width = valueofwidth
        ws.column_dimensions["H"].width = valueofwidth
        ws.column_dimensions["I"].width = valueofwidth
        ws.column_dimensions["J"].width = valueofwidth
        ws.column_dimensions["K"].width = valueofwidth
        ws.column_dimensions["L"].width = valueofwidth
        ws.column_dimensions["M"].width = valueofwidth
        ws.column_dimensions["N"].width = valueofwidth

        # 重置所有計數器並彙總收集的統計信息
        os.popen("adb -s "+ip+" shell dumpsys gfxinfo "+packageName+" reset")
        print "清理幀信息回到初始狀態"

        # 模擬滑動頁面操作
        for i in range (1,3):
            print "執行滑動頁面操作" + str(i) + "次"
            os.system("adb -s "+ip+" shell input swipe 700 2000 700 200") 
            time.sleep(1)

        # 過濾、篩選精確的幀時間信息
        command = "adb -s "+ip+" shell dumpsys gfxinfo "+packageName+" framestats | grep -A 120 'Flags'"
        r = os.popen(command)
        info = r.read().splitlines()

        # 數據處理中
        print "緩存數據中......"
        for line in info:  #按行遍歷
            # line = line.strip('\r\n')
            eachline = line.split(',')
            # 將行寫入Excel表格
            ws.append(eachline)
            # print line

        # 新建sheet用來統計數據
        resultsheet = wb.create_sheet("result",0)
        resultsheet.column_dimensions["A"].width = valueofwidth
        resultsheet.column_dimensions["B"].width = valueofwidth
        resultsheet.column_dimensions["C"].width = valueofwidth
        resultsheet.column_dimensions["D"].width = valueofwidth
        resultsheet.column_dimensions["E"].width = valueofwidth
        resultsheet.column_dimensions["F"].width = valueofwidth
        resultsheet.column_dimensions["G"].width = valueofwidth
        resultsheet.column_dimensions["H"].width = valueofwidth
        resultsheet.column_dimensions["I"].width = valueofwidth

        # 為結果頁添加title說明
        resultsheet.append(titlelist)
        resultsheet.append(subtitlelist)
        # resultsheet.RowDimension(height = 5)

        # 填入公式，cell值由納秒轉換為毫秒
        for i in range(3,123):
            resultsheet.cell(row = i, column = 1, value = "=data!C" + str(i-1) + "-data!B"+ str(i-1))

        for i in range(3,123):
            value = "=(data!G" + str(i-1) + "-data!F"+ str(i-1)
            resultsheet.cell(row = i, column = 2, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!H" + str(i-1) + "-data!G"+ str(i-1)
            resultsheet.cell(row = i, column = 3, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!I" + str(i-1) + "-data!G"+ str(i-1)
            resultsheet.cell(row = i, column = 4, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!K" + str(i-1) + "-data!I"+ str(i-1)
            resultsheet.cell(row = i, column = 5, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!L" + str(i-1) + "-data!K"+ str(i-1)
            resultsheet.cell(row = i, column = 6, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!L" + str(i-1) + "-data!K"+ str(i-1)
            resultsheet.cell(row = i, column = 7, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!N" + str(i-1) + "-data!B"+ str(i-1)
            resultsheet.cell(row = i, column = 8, value = value + ")/1000000")

        # 插入平均值16ms的列
        for i in range(3,123):
            resultsheet.cell(row = i, column = 9, value = 16)

        # 插入平均Frame值
        resultsheet['J1'] = "平均值ms"
        resultsheet['J2'] = "=AVERAGEA(H3:H122)"

        # 畫圖準備
        chart = LineChart()
        chart.title = titlename + str(j)
        # chart.style = 5       #style都很醜，還不如默認的
        chart.y_axis.title = 'ms'
        chart.x_axis.title = 'Frame'
        chart.width = 30
        chart.height = 15

        # data選取範圍
        data = Reference(resultsheet, min_col=8, min_row=2, max_col=9, max_row=122)
        chart.add_data(data, titles_from_data=True)

        # 創建圖表,在B3位置插入
        resultsheet.add_chart(chart,"B3")

        #記錄時間戳作為文件名
        # filename = time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time())) + ".xlsx"
        # wb.save(filename)

        #以執行名稱 titlename作為文件名
        filename2 = titlename + str(j) + ".xlsx"
        wb.save(filename2)

        # 數據完畢
        print "緩存處理完畢，保存數據到本地" + str(filename2)
        time.sleep(3)


import datetime


def dumpsys_gfxinfo(ip,packageName,startIntent,run_times,wb_in=None,reset_sysUI=False):
    #os.system("adb -s "+ip+" root")
    os.system("adb connect "+ip)
    #os.system("adb -s "+ip+" shell setprop debug.hwui.profile true")
    if reset_sysUI:
        os.system("adb -s "+ip+" shell busybox pkill com.android.systemui")
    
    time.sleep(10)

    # start recent app switch thd
    averageSysUIRespTime=[0]
    stopThdEvent = threading.Event()
    appSwitchThd=threading.Thread(target=__openRecentApp,args=(ip,averageSysUIRespTime,stopThdEvent,))
    #appSwitchThd.start()
    titlelist = ['Draw','Prepare','Process','Execute','totalTime','16ms','AverageTime','DropFrameCount']

    # 要測試測模塊名，最後文件會以該名稱命名
    titlename = "FPS-statics"
    print "Starting"
    # os.system("adb -s "+ip+" shell am start "+packageName+"/"+startIntent) 
    if wb_in == None:
        wb = Workbook()
    
        ws = wb.active
        ws.title = "result"
            
        #dinoProcess=doDinoTest(ip)
    else:
        wb = wb_in
        ws = wb.create_sheet("result-"+ip,0)
    lineNums=0
    ws.append(titlelist)
    start_dt=datetime.datetime.now()
    for j in range(1,run_times):
        time.sleep(1)
        print "開始執行第" + str(j) + "遍"

        
        valueofwidth = 16
        ws.column_dimensions["A"].width = valueofwidth
        ws.column_dimensions["B"].width = valueofwidth
        ws.column_dimensions["C"].width = valueofwidth
        ws.column_dimensions["D"].width = valueofwidth
        ws.column_dimensions["E"].width = valueofwidth
        ws.column_dimensions["F"].width = valueofwidth
        ws.column_dimensions["G"].width = valueofwidth
        ws.column_dimensions["H"].width = valueofwidth
        # ws.column_dimensions["I"].width = valueofwidth
        # ws.column_dimensions["J"].width = valueofwidth
        # ws.column_dimensions["K"].width = valueofwidth
        # ws.column_dimensions["L"].width = valueofwidth
        # ws.column_dimensions["M"].width = valueofwidth
        # ws.column_dimensions["N"].width = valueofwidth

        # 重置所有計數器並彙總收集的統計信息
        #os.popen("adb -s "+ip+" shell dumpsys gfxinfo "+packageName+" reset")
        os.popen("adb -s "+ip+" shell dumpsys gfxinfo reset")
        print "清理幀信息回到初始狀態"

        # 模擬滑動頁面操作
        
        # for i in range (1,3):
        #     print "執行滑動頁面操作" + str(i) + "次"
        #     os.system("adb -s "+ip+" shell input swipe 536 700 533 300") 
        #     time.sleep(1)
        #     os.system("adb -s "+ip+" shell input swipe 533 160 536 500") 
        #     time.sleep(1)
        
        # 過濾、篩選精確的幀時間信息
        
        #command = "adb -s "+ip+" shell dumpsys gfxinfo "+packageName+" | grep -A 128 -P 'Prepare\\tProcess'"
        command = "sh -c \"adb -s "+ip+" shell dumpsys gfxinfo me.zhanghai.android.materialprogressbar.sample| grep -A 128 -P 'Prepare\\tProcess'\" 2>&1"
        print command
        r = os.popen(command)
        
        info = r.read().splitlines()
        
        
        # 數據處理中
        #print info
        print "緩存數據中......"
        
        
        for line in info:  #按行遍歷
            if line.find("' not found")!=-1: continue
            # line = line.strip('\r\n')
            eachline = line.split('\t')[1:]
            # print eachline
            if len(eachline)==0 or len(eachline)>4: continue
            if is_number(eachline[0]) and is_number(eachline[1]): 
            # 將行寫入Excel表格
                floats = map(float, eachline)
                ws.append(floats)
                lineNums+=1
            # print line
        # titlelist = ['Draw','Prepare','Process','Execute','totalTime','16ms','AverageTime']
        for i in range(2,2+lineNums):
            ws.cell(row = i, column = 6, value = 16)
            ws.cell(row = i, column = 5, value = "=SUM(A%d:D%d)"%(i,i))
        if j==run_times-1:
            
            end_dt=datetime.datetime.now()
            droppedFrameCnt=droppedFPS(ip,start_dt,end_dt)
            print droppedFrameCnt
            # if dinoProcess is not None:
            #     dinoProcess.terminate()
            # 插入平均Frame值
            ws['G2'] = "=AVERAGEA(E2:E%d)"%(lineNums+1)
            ws['H2'] = droppedFrameCnt
            ws['G3'] = ">16ms count"
            ws['G4'] = "=COUNTIF(E2:E%d,\">16\")"%(lineNums+1)
            ws['H3'] = ">33ms count"
            ws['H4'] = "=COUNTIF(E2:E%d,\">33\")"%(lineNums+1)
            ws['G5'] = "Total count"
            ws['G6'] = "=%d"%lineNums
            stopThdEvent.set()
            print "Stopping Thread.."
            #appSwitchThd.join(60)
            
            
            #print "avg:"+str(averageSysUIRespTime[0])
            #ws['H5'] = "AverageSystemUIResponseTime"
            #ws['H6'] = "=%f"%averageSysUIRespTime[0]
        
            # 畫圖準備
            chart = LineChart()
            chart.title = titlename + str(j)
            # chart.style = 5       #style都很醜，還不如默認的
            chart.y_axis.title = 'ms'
            chart.x_axis.title = 'Frame'
            chart.width = 30
            chart.height = 15

            # data選取範圍
            data = Reference(ws, min_col=5, min_row=1, max_col=6, max_row=2+lineNums)
            chart.add_data(data, titles_from_data=True)

            # 創建圖表,在B3位置插入
            ws.add_chart(chart,"B7")

        #記錄時間戳作為文件名
        # filename = time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time())) + ".xlsx"
        # wb.save(filename)

        #以執行名稱 titlename作為文件名
        
        #time.sleep(3)
    
    if wb_in == None:
        filename2 = ip+'_'+titlename+ ".xlsx"
        wb.save("data/"+filename2)
        print "緩存處理完畢，保存數據到本地" + str(filename2)

    # 數據完畢
    
    print "Stopping Thread.."
    

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False



def droppedFPS(ip,start_dt,end_dt):

    start_dt_str = end_dt.strftime("%m-%d %H:%M:%S")
    adb_command="adb -s "+ip+" logcat -t \'"+start_dt_str+".000\' -v time -s Choreographer"
    print adb_command
    r = os.popen(adb_command)
    info = r.read().splitlines()
    year=start_dt.year
    skippedFrameCnt=0
    for line in info:
        print line
        if line.find("I/Choreographer")!=-1:
            line=line.split()
            datetime_object = datetime.datetime.strptime(str(year)+"-"+line[0]+" "+line[1][:-4], '%Y-%m-%d %H:%M:%S')
            if datetime_object > start_dt:
                if(line[4]=="Skipped"):
                    skippedFrameCnt+=int(line[5])
                else:
                    skippedFrameCnt+=int(line[4])
            if datetime_object < end_dt:
                break
    return skippedFrameCnt

from mock import patch
import unittest
import multiprocessing



def doDinoTest(ip):
    testProcess= multiprocessing.Process(target=__culebraTest,name=ip+'.testProcess',args=(ip,chromeDino.chromeDino,))
    testProcess.start()
    return testProcess

def __culebraTest(ip,testCase):

    sys.argv.append("-s")
    sys.argv.append(ip)
    sys.argv.append("-v")
    devNull = open(os.devnull, 'w')
    with patch("sys.stdout",devNull):
        with patch("sys.stderr",devNull):
            suite = unittest.defaultTestLoader.loadTestsFromTestCase(testCase)
            unittest.TextTestRunner(verbosity=3,stream=sys.stderr).run(suite)
            sys.stderr.flush()
            sys.stdout.flush()



def __openRecentApp(ip,averageSysUIRespTime,stopThdEvent):
    cnt = 0.0
    totalUseTime =0.0
    lines=[]
    while not stopThdEvent.is_set():
        rsp=os.popen("time --format='%E' adb -s "+ip+" shell input keyevent KEYCODE_APP_SWITCH 2>&1")
        lines.append(rsp.read())
        
    for line in lines:
        if line.find(":")==-1 : continue
        sec=float(line[line.find(':')+1:-1])
        minute=float(line[:line.find(':')])
        totalUseTime += minute*60+sec
        cnt+=+1
    averageSysUIRespTime[0] = totalUseTime / cnt
    





    




        

