#!/usr/bin/python

import utilitys
import os
import GPUMonitor
# GPUMonitor.dumpsys_gfxinfo("192.168.0.10","com.android.chrome","com.google.android.apps.chrome.Main -d https://stackoverflow.com/")
#GPUMonitor.dumpsys_gfxinfo("192.168.0.10","com.example.android.recyclerview",".MainActivity")
print "If dumpsys gfxinfo result is null, remember to run the following command:"
print "telnet ...# setprop debug.hwui.profile true"
print "telnet ...# setprop debug.choreographer.skipwarning 1 "
print "telnet ...# setprop ctl.restart surfaceflinger; setprop ctl.restart zygote"
print "After all, you need to reconnect adb."


GPUMonitor.dumpsys_gfxinfo("192.168.3.10","com.android.chrome","com.google.android.apps.chrome.Main -d http://wayou.github.io/t-rex-runner/",100)

exit()

import threading
thd_pool=[]
ip_prefix="192.168.0."
start_ip=50
MAX_NODE_IN_CLUSTER=35
from openpyxl import Workbook

wb= Workbook()
ws = wb.active
ws.title="statics"

def _exit():

    _1stSheetName = wb.worksheets[0].title
    _2ndSheetName = wb.worksheets[-1].title
    # Generate Cluster Statics
    ws['A1']="Avg>16msCnt"
    ws['B1']="=SUMPRODUCT(COUNTIF(INDIRECT(\"'result-192.168.0.\"&ROW(%d:%d)&\"'!G2\"),\">16\"))"%(start_ip,MAX_NODE_IN_CLUSTER+start_ip-1)
    ws['A2']=">100>16msCnt"
    ws['B2']="=SUMPRODUCT(COUNTIF(INDIRECT(\"'result-192.168.0.\"&ROW(%d:%d)&\"'!G4\"),\">100\"))"%(start_ip,MAX_NODE_IN_CLUSTER+start_ip-1)
    ws['A3']=">50>32msCnt"
    ws['B3']="=SUMPRODUCT(COUNTIF(INDIRECT(\"'result-192.168.0.\"&ROW(%d:%d)&\"'!H4\"),\">50\"))"%(start_ip,MAX_NODE_IN_CLUSTER+start_ip-1)

    fileName="ClusterFPS_Statics.xlsx"

    wb.save(fileName)

try:
    for offset in range(0,MAX_NODE_IN_CLUSTER):
        ip=ip_prefix+str(start_ip+offset)
        thd = threading.Thread(target=GPUMonitor.dumpsys_gfxinfo,args=(ip,"","",10,wb,))
        thd.start()
        thd_pool.append(thd)

    import time

    for thd in thd_pool:
        while thd.is_alive():
            time.sleep(0.5)
except KeyboardInterrupt as e:
    _exit()

_exit()












