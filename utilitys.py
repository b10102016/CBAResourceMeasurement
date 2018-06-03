# -*- coding: utf-8 -*-

import telnetlib
import time
import socket

sleep = time.sleep # alias

def deviceShell(cmds,ip,time_out=10):
    tn = telnetlib.Telnet()
    
    try:
        tn = telnetlib.Telnet(ip , timeout=time_out)
        #tn.set_debuglevel(2)
        sleep(0.2)
        for cmd in cmds:
            tn.write(cmd+"\n")
            sleep(0.2)
            tn.read_until(":/ #",timeout=time_out)
        tn.write("exit\n")
        
    except socket.error as e:
        if format(e).find("No route to host"):
            print "No route to host "+ip
            
            return
        else:
            raise e
    finally:
        tn.close()

def deviceShellWithResp(cmd,ip,time_out=10):
    tn = telnetlib.Telnet()
    
    try:
        tn = telnetlib.Telnet(ip , timeout=time_out)
        #tn.set_debuglevel(2)
        response=tn.read_until(":/ #",timeout=time_out)
        sleep(0.2)
        tn.write(cmd+"\n")
        sleep(0.2)
        response=tn.read_until(":/ #",timeout=time_out)
        sleep(0.2)
        tn.write("exit\n")
        if response is None : return ""
        return response
    except socket.error as e:
        if format(e).find("No route to host"):
            print "No route to host "+ip
            return ""
        else:
            raise e
    finally:
        tn.close()
    return ""

def dumpsys_gfxinfo(ip,packageName,startIntent):

    
    # Author：John Hao
    # 測試流暢度平均繪製時長的小腳本
    # 功能：滑動當前頁面兩次，間隔1秒，收集gfxinfo並製表
    # titlename：表示要測試的模塊和最後文件保存的名字
    # |-- 收集gfxinfo framestats詳細幀數據信息
    # |-- 用第三方模塊openpyxl處理收集到的excel數據
    # |-- 輸出2秒內的120幀每一幀的繪製時間
    # |-- 輸出平均繪製時間

    
    import time
    
    import sys
    import os
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles.colors import Color
    from openpyxl.chart import LineChart, Reference, Series

    titlelist = [' UI 線程中發生的工作使其無法及時響應垂直同步信號','應用處理輸入事件所花的時間（ >2 毫秒錶示處理輸入事件時間長）',
                    '正在運行的所有動畫（ObjectAnimator、ViewPropertyAnimator 和通用轉換）所需的時間',
                    '完成佈局和測量階段所需的時間','對樹中的所有視圖調用 View.draw() 所需的時間',
                    '約 > 0.4 毫秒錶示繪製了大量必須上傳到 GPU 的新位圖','GPU 工作量','處理此幀所花的總時間','達標線']
    subtitlelist = ['IntendedVsync','HandleInputStart','AnimationStart','PerformTraversalsStart',
                    'DrawStart','SyncStart','IssueDrawCommandsStart','FrameCompleted','Avg']

    # 要測試測模塊名，最後文件會以該名稱命名
    titlename = "Feed"
    print "Starting"
    os.system("adb -s "+ip+" am start "+packageName+"/"+startIntent) 
    for j in range(1,6):
        time.sleep(1)
        print "開始執行第" + str(j) + "遍"

        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        valueofwidth = 16
        ws.column_dimensions["A"].width = valueofwidth
        ws.column_dimensions["B"].width = valueofwidth
        ws.column_dimensions["C"].width = valueofwidth
        ws.column_dimensions["D"].width = valueofwidth
        ws.column_dimensions["E"].width = valueofwidth
        ws.column_dimensions["F"].width = valueofwidth
        ws.column_dimensions["G"].width = valueofwidth
        ws.column_dimensions["H"].width = valueofwidth
        ws.column_dimensions["I"].width = valueofwidth
        ws.column_dimensions["J"].width = valueofwidth
        ws.column_dimensions["K"].width = valueofwidth
        ws.column_dimensions["L"].width = valueofwidth
        ws.column_dimensions["M"].width = valueofwidth
        ws.column_dimensions["N"].width = valueofwidth

        # 重置所有計數器並彙總收集的統計信息
        os.popen("adb -s "+ip+" shell dumpsys gfxinfo "+packageName+" reset")
        print "清理幀信息回到初始狀態"

        # 模擬滑動頁面操作
        for i in range (1,3):
            print "執行滑動頁面操作" + str(i) + "次"
            os.system("adb -s "+ip+" shell input swipe 700 2000 700 200") 
            time.sleep(1)

        # 過濾、篩選精確的幀時間信息
        command = "adb -s "+ip+" shell dumpsys gfxinfo "+packageName+" framestats | grep -A 120 'Flags'"
        r = os.popen(command)
        info = r.readlines()

        # 數據處理中
        print "緩存數據中......"
        for line in info:  #按行遍歷
            # line = line.strip('\r\n')
            eachline = line.split(',')
            # 將行寫入Excel表格
            ws.append(eachline)
            # print line

        # 新建sheet用來統計數據
        resultsheet = wb.create_sheet("result",0)
        resultsheet.column_dimensions["A"].width = valueofwidth
        resultsheet.column_dimensions["B"].width = valueofwidth
        resultsheet.column_dimensions["C"].width = valueofwidth
        resultsheet.column_dimensions["D"].width = valueofwidth
        resultsheet.column_dimensions["E"].width = valueofwidth
        resultsheet.column_dimensions["F"].width = valueofwidth
        resultsheet.column_dimensions["G"].width = valueofwidth
        resultsheet.column_dimensions["H"].width = valueofwidth
        resultsheet.column_dimensions["I"].width = valueofwidth

        # 為結果頁添加title說明
        resultsheet.append(titlelist)
        resultsheet.append(subtitlelist)
        # resultsheet.RowDimension(height = 5)

        # 填入公式，cell值由納秒轉換為毫秒
        for i in range(3,123):
            resultsheet.cell(row = i, column = 1, value = "=data!C" + str(i-1) + "-data!B"+ str(i-1))

        for i in range(3,123):
            value = "=(data!G" + str(i-1) + "-data!F"+ str(i-1)
            resultsheet.cell(row = i, column = 2, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!H" + str(i-1) + "-data!G"+ str(i-1)
            resultsheet.cell(row = i, column = 3, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!I" + str(i-1) + "-data!G"+ str(i-1)
            resultsheet.cell(row = i, column = 4, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!K" + str(i-1) + "-data!I"+ str(i-1)
            resultsheet.cell(row = i, column = 5, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!L" + str(i-1) + "-data!K"+ str(i-1)
            resultsheet.cell(row = i, column = 6, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!L" + str(i-1) + "-data!K"+ str(i-1)
            resultsheet.cell(row = i, column = 7, value = value + ")/1000000")

        for i in range(3,123):
            value = "=(data!N" + str(i-1) + "-data!B"+ str(i-1)
            resultsheet.cell(row = i, column = 8, value = value + ")/1000000")

        # 插入平均值16ms的列
        for i in range(3,123):
            resultsheet.cell(row = i, column = 9, value = 16)

        # 插入平均Frame值
        resultsheet['J1'] = "平均值ms"
        resultsheet['J2'] = "=AVERAGEA(H3:H122)"

        # 畫圖準備
        chart = LineChart()
        chart.title = titlename + str(j)
        # chart.style = 5       #style都很醜，還不如默認的
        chart.y_axis.title = 'ms'
        chart.x_axis.title = 'Frame'
        chart.width = 30
        chart.height = 15

        # data選取範圍
        data = Reference(resultsheet, min_col=8, min_row=2, max_col=9, max_row=122)
        chart.add_data(data, titles_from_data=True)

        # 創建圖表,在B3位置插入
        resultsheet.add_chart(chart,"B3")

        #記錄時間戳作為文件名
        # filename = time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time())) + ".xlsx"
        # wb.save(filename)

        #以執行名稱 titlename作為文件名
        filename2 = titlename + str(j) + ".xlsx"
        wb.save(filename2)

        # 數據完畢
        print "緩存處理完畢，保存數據到本地" + str(filename2)
        time.sleep(3)


